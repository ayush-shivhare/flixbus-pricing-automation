"""
Flixbus Pricing Flagging System — Full Automation Script
Route: Route 1 | Author: Data & Growth Intern Assignment
Approach: Compare Flixbus WAP against comparable AC Sleeper (2+1) buses
          on same route, same DOJ, within ±2h departure window.
Flag threshold: ±15% deviation from comparable median WAP.
"""

import pandas as pd
import numpy as np
import warnings
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings('ignore')

# ============================================================
# CONFIGURATION
# ============================================================
CONFIG = {
    'flixbus_operator_name': 'Flixbus',
    'target_bus_types': ['A/C Sleeper (2+1)', 'AC Sleeper (2+1)', 'A/C Seater / Sleeper (2+1)',
                         'A/C Seater/Sleeper (2+1)', 'Bharat Benz A/C Sleeper (2+1)',
                         'Volvo Multi-Axle A/C Sleeper (2+1)', 'Volvo 9600 Multi-Axle A/C Sleeper (2+1)',
                         'Volvo B11R Multi-Axle A/C Sleeper (2+1)'],
    'flag_threshold_pct': 15,          # % above/below comparable median to raise flag
    'min_comparables': 3,              # Minimum comparables required for reliable flagging
    'departure_window_hours': 2,       # ±hours window for comparable matching
    'daytime_cutoff_hour': 17,         # Before this hour = daytime (lower price adjustment)
    'daytime_discount': 0.92,          # Day buses priced ~8% lower than evening
    'seater_sleeper_discount': 0.93,   # Mixed config buses priced ~7% lower than pure sleeper
    'non_ac_exclude': True,            # Exclude non-AC buses from comparables
    'min_reviews_threshold': 5,        # Minimum reviews for bus to be used as comparable
    'output_file': 'Flixbus_Pricing_Flagging_Output.xlsx',
}


# ============================================================
# STEP 1: LOAD & CLEAN DATA
# ============================================================
def load_data(filepath: str) -> pd.DataFrame:
    """
    Load raw bus data CSV/Excel file.
    Handles German decimal format (comma as decimal separator).
    """
    if filepath.endswith('.xlsx') or filepath.endswith('.xls'):
        df = pd.read_excel(filepath)
    else:
        df = pd.read_csv(filepath, sep='\t', encoding='utf-8')

    # Fix decimal separators (German format uses comma)
    for col in ['Weighted Average Price', 'Total Ratings', 'Bus Score']:
        if col in df.columns:
            df[col] = df[col].astype(str).str.replace(',', '.').str.strip()
            df[col] = pd.to_numeric(df[col], errors='coerce')

    # Parse DOJ
    if 'DOJ' in df.columns:
        df['DOJ'] = pd.to_datetime(df['DOJ'], dayfirst=True, errors='coerce')

    # Parse departure time to hours (numeric) for window comparison
    if 'Departure Time' in df.columns:
        df['Dep_Hour'] = df['Departure Time'].apply(parse_hour)

    # Boolean flags
    for col in ['Is AC', 'Is Sleeper', 'Is Seater']:
        if col in df.columns:
            df[col] = df[col].map({'TRUE': True, 'True': True, True: True,
                                   'FALSE': False, 'False': False, False: False,
                                   'NA': np.nan})

    # Extract rank number for sorting
    if 'SRP Rank' in df.columns:
        df['Rank_Num'] = df['SRP Rank'].astype(str).str.extract(r'^(\d+)').astype(float)

    # Clean number of reviews
    if 'Number of Reviews' in df.columns:
        df['Number of Reviews'] = pd.to_numeric(df['Number of Reviews'], errors='coerce').fillna(0)

    print(f"[LOAD] Loaded {len(df)} rows | {df['DOJ'].nunique()} unique dates | "
          f"{df['Route Number'].nunique() if 'Route Number' in df.columns else 'N/A'} routes")

    return df


def parse_hour(time_str: str) -> float:
    """Convert HH:MM to decimal hours. Handles midnight wraparound."""
    try:
        parts = str(time_str).strip().split(':')
        h, m = int(parts[0]), int(parts[1])
        return h + m / 60
    except Exception:
        return np.nan


# ============================================================
# STEP 2: IDENTIFY SIMILAR BUSES (COMPARABLES)
# ============================================================
def get_comparables(
    flixbus_row: pd.Series,
    all_buses: pd.DataFrame,
    config: dict
) -> pd.DataFrame:
    """
    Find comparable buses for a single Flixbus service.

    Similarity criteria (in order of priority):
    1. Same Route Number
    2. Same DOJ (date of journey)
    3. Not Flixbus (exclude self)
    4. AC bus only
    5. Is Sleeper (or Is Sleeper + Seater for mixed Flixbus)
    6. Departure time within ±window_hours
    7. Minimum review count

    Returns: DataFrame of comparable buses.
    """
    route = flixbus_row.get('Route Number', None)
    doj = flixbus_row.get('DOJ', None)
    dep_hour = flixbus_row.get('Dep_Hour', np.nan)
    window = config['departure_window_hours']
    is_seater_sleeper = (flixbus_row.get('Is Seater', False) is True and
                         flixbus_row.get('Is Sleeper', False) is True)

    mask = (
        (all_buses['Operator'] != config['flixbus_operator_name']) &
        (all_buses['DOJ'] == doj) &
        (all_buses['Is AC'] == True) &
        (all_buses['Is Sleeper'] == True) &
        (all_buses['Number of Reviews'] >= config['min_reviews_threshold'])
    )

    if route:
        mask &= (all_buses['Route Number'] == route)

    # Non-AC exclusion
    if config['non_ac_exclude']:
        mask &= (all_buses['Is AC'] == True)

    comps = all_buses[mask].copy()

    # Filter by departure window (handle midnight wraparound)
    if not np.isnan(dep_hour) and 'Dep_Hour' in comps.columns:
        def in_window(comp_hour):
            if pd.isna(comp_hour):
                return False
            diff = abs(comp_hour - dep_hour)
            # Handle midnight wraparound (e.g., 23:00 and 01:00 are 2h apart, not 22h)
            diff = min(diff, 24 - diff)
            return diff <= window

        comps = comps[comps['Dep_Hour'].apply(in_window)]

    # If too few, expand window
    if len(comps) < config['min_comparables']:
        expanded_mask = (
            (all_buses['Operator'] != config['flixbus_operator_name']) &
            (all_buses['DOJ'] == doj) &
            (all_buses['Is AC'] == True) &
            (all_buses['Is Sleeper'] == True)
        )
        if route:
            expanded_mask &= (all_buses['Route Number'] == route)
        comps = all_buses[expanded_mask].copy()

    return comps


# ============================================================
# STEP 3: CALCULATE REFERENCE PRICE & FLAG
# ============================================================
def calculate_flag(
    flixbus_row: pd.Series,
    comparables: pd.DataFrame,
    config: dict
) -> dict:
    """
    Compare Flixbus WAP against comparable median and flag if deviation > threshold.

    Adjustments:
    - Daytime departure: apply daytime_discount to reference
    - Mixed seater/sleeper: apply seater_sleeper_discount to reference
    """
    flixbus_wap = flixbus_row.get('Weighted Average Price', np.nan)
    dep_hour = flixbus_row.get('Dep_Hour', np.nan)
    bus_type = str(flixbus_row.get('Bus Type', ''))
    is_seater_sleeper = ('Seater' in bus_type and 'Sleeper' in bus_type)
    is_daytime = (not pd.isna(dep_hour) and dep_hour < config['daytime_cutoff_hour']
                  and dep_hour >= 5)  # 05:00–17:00 = daytime

    if len(comparables) < config['min_comparables']:
        return _no_flag_result(flixbus_wap, len(comparables), 'INSUFFICIENT_DATA')

    comp_waps = comparables['Weighted Average Price'].dropna()
    if len(comp_waps) == 0:
        return _no_flag_result(flixbus_wap, 0, 'NO_COMP_PRICE_DATA')

    # Reference price = median of comparable WAPs
    ref_median = comp_waps.median()
    ref_mean = comp_waps.mean()
    ref_p25 = comp_waps.quantile(0.25)
    ref_p75 = comp_waps.quantile(0.75)
    n_comps = len(comp_waps)

    # Apply adjustments
    adj_factor = 1.0
    adjustments_applied = []
    if is_daytime:
        adj_factor *= config['daytime_discount']
        adjustments_applied.append(f'Daytime ×{config["daytime_discount"]}')
    if is_seater_sleeper:
        adj_factor *= config['seater_sleeper_discount']
        adjustments_applied.append(f'Seater/Sleeper ×{config["seater_sleeper_discount"]}')

    adj_ref_median = ref_median * adj_factor
    adj_ref_mean = ref_mean * adj_factor
    adj_ref_p25 = ref_p25 * adj_factor
    adj_ref_p75 = ref_p75 * adj_factor

    if pd.isna(flixbus_wap) or adj_ref_median == 0:
        return _no_flag_result(flixbus_wap, n_comps, 'INVALID_PRICE')

    diff = flixbus_wap - adj_ref_median
    diff_pct = (diff / adj_ref_median) * 100
    threshold = config['flag_threshold_pct']

    if diff_pct > threshold:
        flag = 'TOO HIGH'
    elif diff_pct < -threshold:
        flag = 'TOO LOW'
    else:
        flag = 'OK'

    # Build comparable bus names list
    comp_names = comparables[['Operator', 'Weighted Average Price']].copy()
    comp_names = comp_names.dropna(subset=['Weighted Average Price'])
    comp_names = comp_names.sort_values('Weighted Average Price')
    comp_str = ' | '.join([f"{r['Operator']} (₹{r['Weighted Average Price']:.0f})"
                            for _, r in comp_names.head(8).iterrows()])

    return {
        'Flag': flag,
        'Flixbus WAP (₹)': round(flixbus_wap, 0),
        'Comp. Median (₹)': round(adj_ref_median, 0),
        'Comp. Mean (₹)': round(adj_ref_mean, 0),
        'Comp. P25 (₹)': round(adj_ref_p25, 0),
        'Comp. P75 (₹)': round(adj_ref_p75, 0),
        'N Comparables': n_comps,
        'Price Diff (₹)': round(diff, 0),
        'Price Diff (%)': round(diff_pct, 2),
        'Magnitude (₹)': abs(round(diff, 0)),
        'Adjustments Applied': ', '.join(adjustments_applied) or 'None',
        'Comparable Buses': comp_str,
        'Data Status': 'OK',
    }


def _no_flag_result(flixbus_wap, n_comps, reason):
    return {
        'Flag': 'SKIP', 'Flixbus WAP (₹)': flixbus_wap,
        'Comp. Median (₹)': np.nan, 'Comp. Mean (₹)': np.nan,
        'Comp. P25 (₹)': np.nan, 'Comp. P75 (₹)': np.nan,
        'N Comparables': n_comps, 'Price Diff (₹)': np.nan,
        'Price Diff (%)': np.nan, 'Magnitude (₹)': np.nan,
        'Adjustments Applied': 'N/A', 'Comparable Buses': 'N/A',
        'Data Status': reason,
    }


# ============================================================
# STEP 4: RUN FULL ANALYSIS
# ============================================================
def run_pricing_analysis(df: pd.DataFrame, config: dict) -> pd.DataFrame:
    """Run flagging analysis for all Flixbus buses in the dataset."""
    flixbus_df = df[df['Operator'] == config['flixbus_operator_name']].copy()
    all_buses = df.copy()

    print(f"[ANALYSIS] Analyzing {len(flixbus_df)} Flixbus services...")
    results = []

    for idx, row in flixbus_df.iterrows():
        comps = get_comparables(row, all_buses, config)
        flag_result = calculate_flag(row, comps, config)

        output_row = {
            'Route': row.get('Route Number', ''),
            'DOJ': row.get('DOJ', '').strftime('%d.%m.%Y') if pd.notna(row.get('DOJ')) else '',
            'SRP Rank': row.get('SRP Rank', ''),
            'Departure Time': row.get('Departure Time', ''),
            'Arrival Time': row.get('Arrival Time', ''),
            'Journey Duration (Min)': row.get('Journey Duration (Min)', ''),
            'Bus Type': row.get('Bus Type', ''),
            'Is AC': row.get('Is AC', ''),
            'Is Sleeper': row.get('Is Sleeper', ''),
            **flag_result,
        }
        results.append(output_row)

    results_df = pd.DataFrame(results)
    print(f"[ANALYSIS] Complete. Flags: {results_df['Flag'].value_counts().to_dict()}")
    return results_df


# ============================================================
# STEP 5: EXPORT TO EXCEL
# ============================================================
def export_to_excel(results_df: pd.DataFrame, config: dict):
    """Write flagging results to a formatted Excel workbook."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Flagging Output"

    # Styles
    RED_FILL = PatternFill("solid", fgColor="FF4444")
    BLUE_FILL = PatternFill("solid", fgColor="4472C4")
    GREEN_FILL = PatternFill("solid", fgColor="70AD47")
    GREY_FILL = PatternFill("solid", fgColor="BFBFBF")
    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    WHITE_BOLD = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    STD = Font(name='Arial', size=10)
    BORDER = Border(
        left=Side(style='thin', color='D0D0D0'),
        right=Side(style='thin', color='D0D0D0'),
        top=Side(style='thin', color='D0D0D0'),
        bottom=Side(style='thin', color='D0D0D0')
    )

    # Title
    ws.merge_cells('A1:Q1')
    ws['A1'] = f'Flixbus Pricing Flagging Report — Generated {datetime.now().strftime("%d %b %Y %H:%M")}'
    ws['A1'].font = Font(name='Arial', bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = HEADER_FILL
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 26

    headers = list(results_df.columns)
    for col, h in enumerate(headers, 1):
        c = ws.cell(2, col, h)
        c.font = WHITE_BOLD
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[2].height = 32

    flag_col_idx = headers.index('Flag') + 1

    for i, row in results_df.iterrows():
        er = i + 3
        flag = str(row.get('Flag', ''))
        if flag == 'TOO HIGH':
            rfill = PatternFill("solid", fgColor="FFCCCC")
            ffill = RED_FILL
            ftext = '⬆ TOO HIGH'
        elif flag == 'TOO LOW':
            rfill = PatternFill("solid", fgColor="CCE5FF")
            ffill = BLUE_FILL
            ftext = '⬇ TOO LOW'
        elif flag == 'OK':
            rfill = PatternFill("solid", fgColor="E8F5E9" if i % 2 == 0 else "FFFFFF")
            ffill = GREEN_FILL
            ftext = '✓ OK'
        else:
            rfill = PatternFill("solid", fgColor="F5F5F5")
            ffill = GREY_FILL
            ftext = flag

        for col, h in enumerate(headers, 1):
            val = row.get(h, '')
            if h == 'Flag':
                val = ftext
            cell = ws.cell(er, col, val)
            cell.font = STD
            cell.fill = ffill if col == flag_col_idx else rfill
            cell.border = BORDER
            cell.alignment = Alignment(horizontal='center', vertical='center',
                                       wrap_text=(col in [flag_col_idx, len(headers)]))
            if col == flag_col_idx:
                cell.font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
            if '(₹)' in h or '(Min)' in h:
                cell.number_format = '#,##0'
            elif '(%)' in h:
                cell.number_format = '+0.00%;-0.00%'
                if isinstance(val, (int, float)) and not pd.isna(val):
                    cell.value = val / 100

        ws.row_dimensions[er].height = 17

    # Column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions[get_column_letter(len(headers))].width = 60  # Comparable buses col
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['C'].width = 10

    ws.freeze_panes = 'A3'
    wb.save(config['output_file'])
    print(f"[EXPORT] Results saved to: {config['output_file']}")


# ============================================================
# STEP 6: SUMMARY REPORT TO CONSOLE
# ============================================================
def print_summary(results_df: pd.DataFrame):
    print("\n" + "="*60)
    print("FLIXBUS PRICING FLAG SUMMARY")
    print("="*60)
    valid = results_df[results_df['Flag'].isin(['TOO HIGH', 'TOO LOW', 'OK'])]
    print(f"Total services analyzed:   {len(results_df)}")
    print(f"  ↑ TOO HIGH (overpriced): {(results_df['Flag']=='TOO HIGH').sum()}")
    print(f"  ↓ TOO LOW (underpriced): {(results_df['Flag']=='TOO LOW').sum()}")
    print(f"  ✓ OK (within threshold): {(results_df['Flag']=='OK').sum()}")
    print(f"  ⚠ Skipped (insufficient data): {(results_df['Flag']=='SKIP').sum()}")

    flagged = results_df[results_df['Flag'].isin(['TOO HIGH', 'TOO LOW'])]
    if len(flagged):
        print(f"\nTop 5 largest price deviations:")
        top = flagged.nlargest(5, 'Magnitude (₹)')[
            ['DOJ', 'SRP Rank', 'Departure Time', 'Flixbus WAP (₹)',
             'Comp. Median (₹)', 'Price Diff (%)', 'Flag']
        ]
        print(top.to_string(index=False))
    print("="*60 + "\n")


# ============================================================
# MAIN ENTRY POINT
# ============================================================
def main(input_file: str = 'bus_data.csv'):
    """
    Main pipeline:
    1. Load raw data file
    2. Find comparable buses per Flixbus service
    3. Calculate price flags
    4. Export to Excel
    5. Print console summary
    """
    print(f"\n[START] Flixbus Pricing Flagging System")
    print(f"[CONFIG] Flag threshold: ±{CONFIG['flag_threshold_pct']}%")
    print(f"[CONFIG] Departure window: ±{CONFIG['departure_window_hours']}h")
    print(f"[CONFIG] Input file: {input_file}\n")

    df = load_data(input_file)
    results = run_pricing_analysis(df, CONFIG)
    export_to_excel(results, CONFIG)
    print_summary(results)

    return results


if __name__ == '__main__':
    import sys
    input_file = sys.argv[1] if len(sys.argv) > 1 else 'bus_data.csv'
    main(input_file)


# ============================================================
# USAGE EXAMPLE
# ============================================================
"""
HOW TO RUN:
-----------
1. Install dependencies:
   pip install pandas numpy openpyxl

2. Prepare your data:
   - Save the full dataset as a TSV or Excel file
   - Required columns: Route Number, DOJ, Operator, Bus Type, Is AC, Is Seater,
     Is Sleeper, Departure Time, Weighted Average Price, Number of Reviews

3. Run the script:
   python flixbus_pricing_automation.py bus_data.csv

4. Output:
   - Flixbus_Pricing_Flagging_Output.xlsx (Excel report with color-coded flags)
   - Console summary of all flags

SCHEDULING (CRON JOB):
-----------------------
To run automatically every night at midnight:
  0 0 * * * /usr/bin/python3 /path/to/flixbus_pricing_automation.py /path/to/data.csv

INTEGRATION WITH DATA PIPELINE:
---------------------------------
# In your ETL/data pipeline script:
from flixbus_pricing_automation import load_data, run_pricing_analysis, CONFIG
df = load_data('your_data_file.csv')
results = run_pricing_analysis(df, CONFIG)
# Send results to Slack, email, dashboard, etc.
"""
